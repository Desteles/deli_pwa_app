#Основной код

import sqlite3
import logging
import openpyxl
from telegram import error as telegram_error
from datetime import datetime, timedelta
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, KeyboardButton, ReplyKeyboardMarkup, CallbackQuery
from telegram.ext import Application, CommandHandler, CallbackQueryHandler, ContextTypes, ConversationHandler, MessageHandler, filters

# Настройка логирования
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Состояния
# Состояния
SUPPLIER, PAYER, INVOICE, PICKUP, DELIVERY, CARGO_INFO = range(6)
EDITING_FIELD = 6  # Состояние для ожидания ввода нового значения поля
DRIVER_COMPLETED_COUNT = 0


# Подключение к БД
conn = sqlite3.connect('logistics.db', check_same_thread=False)

# === РУЧНЫЕ ДАННЫЕ (вместо регистрации через бот) ===
# Список менеджеров (user_id → имя)
MANAGERS = {
    597890387: "Станислав",   
    2002784191: "Вадим"
}

# Список водителей (user_id → имя)
DRIVERS = {
    8293490412: "Стас Тест"
}

# === КЛАВИАТУРЫ ===
def get_manager_menu():
    keyboard = [
        [InlineKeyboardButton("Запланировать доставку", callback_data="add_delivery")],
        [InlineKeyboardButton("Не выполненные", callback_data="planned_deliveries")],
        [InlineKeyboardButton("Приняты в работу", callback_data="in_work_deliveries")],
        [InlineKeyboardButton("Выполненные", callback_data="completed_deliveries")],
        [InlineKeyboardButton("Скачать таблицу", callback_data="download_table")]
    ]
    return InlineKeyboardMarkup(keyboard)

def get_driver_menu():
    keyboard = [
        [InlineKeyboardButton("Мои доставки", callback_data="driver_deliveries")],
        [InlineKeyboardButton("Посмотреть выполненные", callback_data="driver_completed")],
        [InlineKeyboardButton("Скачать таблицу", callback_data="download_table_driver")]  
    ]
    return InlineKeyboardMarkup(keyboard)


async def driver_deliveries(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    try:
        await query.delete_message()
    except telegram_error.BadRequest as e:
        if "Message to delete not found" in str(e):
            logger.warning("Сообщение уже удалено или не найдено.")
        else:
            raise e

    driver_id = update.effective_user.id

    # Очищаем ВСЕ сохранённые ID сообщений (включая назначенные доставки)
    for key in ['assigned_delivery_messages', 'bot_messages']:
        if key in context.user_data:
            for msg_id in context.user_data[key]:
                try:
                    await context.bot.delete_message(
                        chat_id=query.message.chat_id,
                        message_id=msg_id
                    )
                except Exception as e:
                    logger.error(f"Ошибка при удалении сообщения {msg_id}: {e}")
            context.user_data.pop(key, None)

    cursor = conn.execute(
        """SELECT id, supplier, payer, invoice_number, pickup_address,
                   delivery_address, cargo_info, author_name, status
           FROM deliveries
           WHERE driver_id = ? AND status NOT IN ('доставлено', 'отменено')""",
        (driver_id,)
    )
    deliveries = cursor.fetchall()

    if not deliveries:
        msg = await context.bot.send_message(
            chat_id=query.message.chat_id,
            text="У вас нет активных доставок.",
            reply_markup=get_driver_menu()
        )
        context.user_data['bot_messages'] = [msg.message_id]
        return

    message_ids = []

    for d in deliveries:
        delivery_id = d[0]
        supplier = d[1]
        payer = d[2]
        invoice_number = d[3]
        pickup_addr = d[4]
        delivery_addr = d[5]
        cargo_info = d[6]
        author_name = d[7]
        status = d[8]

        message_text = (
            f"🚚 <b>Доставка №{delivery_id}</b>\n\n"
            f"<b>Поставщик:</b> {supplier}\n"
            f"<b>Плательщик:</b> {payer}\n"
            f"<b>Счёт:</b> {invoice_number}\n"
            f"<b>Адрес загрузки:</b> {pickup_addr or '—'}\n"
            f"<b>Адрес отгрузки:</b> {delivery_addr or '—'}\n"
            f"<b>Габариты/вес/комментарий:</b> {cargo_info or '—'}\n"
            f"<b>Автор заявки:</b> {author_name}\n\n"
        )

        if status == 'принята в работу':
            keyboard = [
                [InlineKeyboardButton(
                    "Доставлено", callback_data=f"delivered_{delivery_id}_yes"
                )]
            ]
        else:
            keyboard = [
                [InlineKeyboardButton(
                    "Принять в работу", callback_data=f"accept_{delivery_id}"
                )]
            ]

        reply_markup = InlineKeyboardMarkup(keyboard)

        msg = await context.bot.send_message(
            chat_id=query.message.chat_id,
            text=message_text,
            reply_markup=reply_markup,
            parse_mode="HTML"
        )
        message_ids.append(msg.message_id)

    # Сохраняем ID всех сообщений с доставками
    context.user_data['bot_messages'] = message_ids

    final_msg = await context.bot.send_message(
        chat_id=query.message.chat_id,
        text="Выберите доставку для работы:",
        reply_markup=get_driver_menu()
    )
    context.user_data['bot_messages'].append(final_msg.message_id)



# === ОБРАБОТЧИКИ ===
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.clear()
    user_id = update.effective_user.id

    # Проверка на менеджера
    if user_id in MANAGERS:
        await update.message.reply_text(
            f"Добро пожаловать, {MANAGERS[user_id]}!",
            reply_markup=get_manager_menu()
        )
        return

    # Проверка на водителя
    if user_id in DRIVERS:
        await update.message.reply_text(
            f"Добро пожаловать, {DRIVERS[user_id]}!",
            reply_markup=get_driver_menu()
        )
        return

    # Если пользователь не в списках — сообщение об ошибке
    await update.message.reply_text(
        "Вы не авторизованы как менеджер или водитель. Обратитесь к администратору."
    )


async def add_delivery(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    # Удаляем предыдущее меню
    await query.delete_message()
    await delete_excel_message(context, update.effective_chat.id)

    # 1. Сначала удаляем сообщения с доставками (если есть)
    if 'delivery_messages' in context.user_data:
        for msg_id in context.user_data['delivery_messages']:
            try:
                await context.bot.delete_message(
                    chat_id=query.message.chat_id,
                    message_id=msg_id
                )
            except Exception as e:
                logger.warning(f"Не удалось удалить сообщение {msg_id}: {e}")
        context.user_data.pop('delivery_messages', None)  # Удаляем ключ

    # 2. Теперь очищаем весь контекст
    context.user_data.clear()
    context.user_data['author_id'] = update.effective_user.id

    # Отправляем первый запрос и сохраняем ID
    msg = await context.bot.send_message(
        chat_id=query.message.chat_id,
        text="Поставщик (обязательно):"
    )
    context.user_data['last_bot_msg_id'] = msg.message_id
    return SUPPLIER


async def supplier_step(update: Update, context: ContextTypes.DEFAULT_TYPE):
    logger.info("supplier_step вызван") 
    if not update.message or not update.message.text:
        logger.error("Нет текста в сообщении")
        return SUPPLIER
    try:
        await context.bot.delete_message(
            chat_id=update.effective_chat.id,
            message_id=update.message.message_id
        )
    except:
        pass

    if 'last_bot_msg_id' in context.user_data:
        try:
            await context.bot.delete_message(
                chat_id=update.effective_chat.id,
                message_id=context.user_data['last_bot_msg_id']
            )
        except:
            pass

    supplier = update.message.text.strip()
    if not supplier:
        msg = await update.message.reply_text(
            "Поставщик не может быть пустым. Введите ещё раз:",
        )
        context.user_data['last_bot_msg_id'] = msg.message_id
        return SUPPLIER

    context.user_data['supplier'] = supplier

    msg = await update.message.reply_text(
        "Плательщик (обязательно):",
    )
    context.user_data['last_bot_msg_id'] = msg.message_id
    return PAYER

async def payer_step(update: Update, context: ContextTypes.DEFAULT_TYPE):
    logger.info("payer_step вызван")
    try:
        await context.bot.delete_message(
            chat_id=update.effective_chat.id,
            message_id=update.message.message_id
        )
    except:
        pass

    if 'last_bot_msg_id' in context.user_data:
        try:
            await context.bot.delete_message(
                chat_id=update.effective_chat.id,
                message_id=context.user_data['last_bot_msg_id']
            )
        except:
            pass

    payer = update.message.text.strip()
    if not payer:
        msg = await update.message.reply_text(
            "Плательщик не может быть пустым. Введите ещё раз:",
        )
        context.user_data['last_bot_msg_id'] = msg.message_id
        return PAYER

    context.user_data['payer'] = payer
    msg = await update.message.reply_text(
        "Номер счёта (обязательно):",
    )
    context.user_data['last_bot_msg_id'] = msg.message_id
    return INVOICE


async def invoice_step(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        await context.bot.delete_message(
            chat_id=update.effective_chat.id,
            message_id=update.message.message_id
        )
    except:
        pass

    if 'last_bot_msg_id' in context.user_data:
        try:
            await context.bot.delete_message(
                chat_id=update.effective_chat.id,
                message_id=context.user_data['last_bot_msg_id']
            )
        except:
            pass

    invoice = update.message.text.strip()
    if not invoice:
        msg = await update.message.reply_text(
            "Номер счёта не может быть пустым. Введите ещё раз:",
        )
        context.user_data['last_bot_msg_id'] = msg.message_id
        return INVOICE

    context.user_data['invoice'] = invoice
    msg = await update.message.reply_text(
        "Адрес загрузки (опционально):",
    )
    context.user_data['last_bot_msg_id'] = msg.message_id
    return PICKUP


async def pickup_step(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        await context.bot.delete_message(
            chat_id=update.effective_chat.id,
            message_id=update.message.message_id
        )
    except:
        pass

    if 'last_bot_msg_id' in context.user_data:
        try:
            await context.bot.delete_message(
                chat_id=update.effective_chat.id,
                message_id=context.user_data['last_bot_msg_id']
            )
        except:
            pass

    pickup = update.message.text.strip() or None
    context.user_data['pickup'] = pickup
    msg = await update.message.reply_text(
        "Адрес отгрузки:",
    )
    context.user_data['last_bot_msg_id'] = msg.message_id
    return DELIVERY


async def delivery_step(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        await context.bot.delete_message(
            chat_id=update.effective_chat.id,
            message_id=update.message.message_id
        )
    except:
        pass

    if 'last_bot_msg_id' in context.user_data:
        try:
            await context.bot.delete_message(
                chat_id=update.effective_chat.id,
                message_id=context.user_data['last_bot_msg_id']
            )
        except:
            pass

    delivery = update.message.text.strip() or None
    context.user_data['delivery'] = delivery
    msg = await update.message.reply_text(
        "Габариты/вес/комментарий:",
    )
    context.user_data['last_bot_msg_id'] = msg.message_id
    return CARGO_INFO


async def cargo_info_step(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        await context.bot.delete_message(
            chat_id=update.effective_chat.id,
            message_id=update.message.message_id
        )
    except:
        pass

    if 'last_bot_msg_id' in context.user_data:
        try:
            await context.bot.delete_message(
                chat_id=update.effective_chat.id,
                message_id=context.user_data['last_bot_msg_id']
            )
        except:
            pass

    cargo_info = update.message.text.strip() or None
    context.user_data['cargo_info'] = cargo_info

    # Дальше — ваш существующий код сохранения в БД и отправки итогового сообщения
    author_id = context.user_data['author_id']
    author_name = MANAGERS.get(author_id)

    if not author_name:
        await update.message.reply_text("Ошибка: вы не авторизованы как менеджер.")
        context.user_data.clear()
        return ConversationHandler.END

    try:
        cursor = conn.execute(
            """INSERT INTO deliveries
            (supplier, payer, invoice_number, pickup_address, delivery_address,
             cargo_info, status, author_name)
            VALUES (?, ?, ?, ?, ?, ?, 'черновик', ?)""",
            (
                context.user_data['supplier'],
                context.user_data['payer'],
                context.user_data['invoice'],
                context.user_data['pickup'],
                context.user_data['delivery'],
                context.user_data['cargo_info'],
                author_name
            )
        )
        delivery_id = cursor.lastrowid
        conn.commit()

        msg = (f"✅ Доставка создана! ID: {delivery_id}\n"
               f"Поставщик: {context.user_data['supplier']}\n"
               f"Плательщик: {context.user_data['payer']}\n"
               f"Номер счёта: {context.user_data['invoice']}\n"
               f"Адрес загрузки: {context.user_data['pickup'] or '—'}\n"
               f"Адрес отгрузки: {context.user_data['delivery'] or '—'}\n"
               f"Габариты/вес: {context.user_data['cargo_info'] or '—'}\n"
               f"Автор: {author_name}\n"
               f"Статус: черновик")

        # Отправляем итоговое сообщение (без удаления, т. к. это финальный результат)
        await update.message.reply_text(
            msg,
            reply_markup=get_manager_menu()
        )
        context.user_data.clear()
        return ConversationHandler.END

    except sqlite3.Error as e:
        logger.error(f"Ошибка при сохранении доставки: {e}")
        await update.message.reply_text("Произошла ошибка при сохранении доставки. Попробуйте ещё раз.")
        return ConversationHandler.END




# Обработка невыполненных доставок
async def planned_deliveries(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    # Удаляем старое меню
    await query.delete_message()
    await delete_excel_message(context, query.message.chat_id)

    # Очищаем список сообщений с доставками (если был предыдущий вызов)
    if 'delivery_messages' in context.user_data:
        for msg_id in context.user_data['delivery_messages']:
            try:
                await context.bot.delete_message(
                    chat_id=query.message.chat_id,
                    message_id=msg_id
                )
            except:
                pass
        context.user_data.pop('delivery_messages', None)

    cursor = conn.execute(
        """SELECT id, supplier, payer, invoice_number, pickup_address,
                   delivery_address, cargo_info, author_name
           FROM deliveries WHERE status='черновик'"""
    )
    deliveries = cursor.fetchall()

    if not deliveries:
        msg = await context.bot.send_message(
            chat_id=query.message.chat_id,
            text="Нет незапланированных доставок.",
            reply_markup=get_manager_menu()
        )
        context.user_data['delivery_messages'] = [msg.message_id]
        return

    message_ids = []
    for d in deliveries:
        msg_text = (f"📦 Доставка №{d[0]}\n"
                    f"Поставщик: {d[1]}\n"
                    f"Плательщик: {d[2]}\n"
                    f"Счёт: {d[3]}\n"
                    f"Адрес загрузки: {d[4] or '—'}\n"
                    f"Адрес отгрузки: {d[5] or '—'}\n"
                    f"Габариты/вес: {d[6] or '—'}\n"
                    f"Автор: {d[7]}\n"
                    f"Статус: черновик\n")

        keyboard = [
            [InlineKeyboardButton("Назначить водителя", callback_data=f"assign_driver_{d[0]}")],
            [InlineKeyboardButton("Редактировать доставку", callback_data=f"edit_delivery_{d[0]}")]
        ]

        msg = await context.bot.send_message(
            chat_id=query.message.chat_id,
            text=msg_text,
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
        message_ids.append(msg.message_id)

    # Сохраняем все ID сообщений с доставками
    context.user_data['delivery_messages'] = message_ids


    # Сообщение с инструкцией
    msg = await context.bot.send_message(
        chat_id=query.message.chat_id,
        text="Выберите действие:",
        reply_markup=get_manager_menu()
    )
    context.user_data['delivery_messages'].append(msg.message_id)


# Обработчик редактировать доставку
async def edit_delivery(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    # Удаляем предыдущее сообщение (меню или список доставок)
    await query.delete_message()

    delivery_id = int(query.data.split("_")[2])
    context.user_data['edit_delivery_id'] = delivery_id

    # Очищаем все сообщения с доставками (если были)
    if 'delivery_messages' in context.user_data:
        for msg_id in context.user_data['delivery_messages']:
            try:
                await context.bot.delete_message(
                    chat_id=query.message.chat_id,
                    message_id=msg_id
                )
            except:
                pass
        context.user_data.pop('delivery_messages', None)

    # Получаем текущие данные
    cursor = conn.execute(
        "SELECT supplier, payer, invoice_number, pickup_address, delivery_address, cargo_info "
        "FROM deliveries WHERE id=?", (delivery_id,)
    )
    row = cursor.fetchone()
    if not row:
        await query.edit_message_text("Доставка не найдена.")
        return

    # Формируем меню выбора поля
    fields = {
        "supplier": "Поставщик",
        "payer": "Плательщик",
        "invoice_number": "Номер счёта",
        "pickup_address": "Адрес загрузки",
        "delivery_address": "Адрес отгрузки",
        "cargo_info": "Комментарий о грузе"
    }
    keyboard = []
    for key, label in fields.items():
        keyboard.append([
            InlineKeyboardButton(
                f"Изменить {label}", callback_data=f"field:{key}:{delivery_id}"
            )
        ])

    # Добавляем кнопку "Назад в меню"
    keyboard.append([InlineKeyboardButton("Назад в меню", callback_data="back_to_menu")])

    msg = (f"Редактирование доставки №{delivery_id}\n\n"
           f"Поставщик: {row[0]}\n"
           f"Плательщик: {row[1]}\n"
           f"Счёт: {row[2]}\n"
           f"Адрес загрузки: {row[3] or '—'}\n"
           f"Адрес отгрузки: {row[4] or '—'}\n"
           f"Комментарий: {row[5] or '—'}")

    await context.bot.send_message(
        chat_id=query.message.chat_id,
        text=msg,
        reply_markup=InlineKeyboardMarkup(keyboard)
    )

# Обработчик выбора поля
async def select_edit_field(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    await query.delete_message()

    parts = query.data.split(":", 2)
    if len(parts) != 3:
        await query.edit_message_text("Ошибка: некорректный callback_data.")
        return

    try:
        delivery_id = int(parts[2])
    except ValueError:
        await query.edit_message_text("Ошибка: ID доставки должен быть числом.")
        return

    field = parts[1]

    valid_fields = ["supplier", "payer", "invoice_number", "pickup_address", "delivery_address", "cargo_info"]
    if field not in valid_fields:
        await query.edit_message_text("Ошибка: неизвестное поле для редактирования.")
        return

    context.user_data['edit_field'] = field
    context.user_data['edit_id'] = delivery_id

    cursor = conn.execute(f"SELECT {field} FROM deliveries WHERE id=?", (delivery_id,))
    current_value = cursor.fetchone()
    if not current_value:
        await query.edit_message_text("Ошибка: доставка не найдена.")
        return

    current_value = current_value[0] or "—"

    field_names = {
        "supplier": "поставщика",
        "payer": "плательщика",
        "invoice_number": "номера счёта",
        "pickup_address": "адреса загрузки",
        "delivery_address": "адреса отгрузки",
        "cargo_info": "комментария о грузе"
    }

    # Отправляем сообщение и сохраняем его ID
    msg = await context.bot.send_message(
        chat_id=update.effective_chat.id,
        text=(f"Текущее значение: {current_value}\n\n"
              f"Введите новое значение для {field_names[field]}:"),
    )
    context.user_data['last_bot_msg_id'] = msg.message_id  # Сохраняем ID

    return EDITING_FIELD


#Сохранение изменений
async def save_edited_field(update: Update, context: ContextTypes.DEFAULT_TYPE):
    logger.info(f"[save_edited_field] Получен ввод: {update.message.text}")
    logger.info(f"[save_edited_field] context.user_data: {context.user_data}")

    if not context.user_data.get('edit_field') or not context.user_data.get('edit_id'):
        logger.warning("[save_edited_field] Отсутствуют edit_field или edit_id в context.user_data")
        await update.message.reply_text(
            "Ошибка: не удалось определить поле для редактирования. Вернитесь в меню.",
            reply_markup=get_manager_menu()
        )
        context.user_data.clear()
        return ConversationHandler.END

    new_value = update.message.text.strip()
    field = context.user_data['edit_field']
    delivery_id = context.user_data['edit_id']

    logger.info(f"[save_edited_field] Попытка обновить: delivery_id={delivery_id}, field={field}, new_value={new_value}")


    try:
        conn.execute(f"UPDATE deliveries SET {field}=? WHERE id=?", (new_value, delivery_id))
        conn.commit()
        
        logger.info(f"[save_edited_field] Поле {field} успешно обновлено для delivery_id={delivery_id}")

        # Удаляем сообщение бота с запросом
        if 'last_bot_msg_id' in context.user_data:
            try:
                await context.bot.delete_message(
                    chat_id=update.effective_chat.id,
                    message_id=context.user_data['last_bot_msg_id']
                )
            except Exception as e:
                logger.warning(f"Не удалось удалить сообщение бота: {e}")

        # Удаляем сообщение пользователя с введённым значением
        try:
            await context.bot.delete_message(
                chat_id=update.effective_chat.id,
                message_id=update.message.message_id
            )
        except Exception as e:
            logger.warning(f"Не удалось удалить сообщение пользователя: {e}")

        # Словарь для отображения понятных названий полей
        field_names = {
            "supplier": "Поставщик",
            "payer": "Плательщик",
            "invoice_number": "Номер счёта",
            "pickup_address": "Адрес загрузки",
            "delivery_address": "Адрес отгрузки",
            "cargo_info": "Комментарий о грузе"
        }
        display_name = field_names.get(field, field)  # Если поля нет в словаре — используем ключ

        # Формируем итоговое сообщение с новым значением
        await update.message.reply_text(
            f"✅ Значение поля «{display_name}» успешно изменено на «{new_value}»!\n\n"
            "Возвращаемся в главное меню...",
            reply_markup=get_manager_menu()
        )

    except sqlite3.Error as e:
        logger.error(f"[save_edited_field] Ошибка SQL при обновлении поля {field}: {e}")
        await update.message.reply_text(
            "❌ Произошла ошибка при сохранении изменений. Попробуйте ещё раз.",
            reply_markup=get_manager_menu()
        )
        return

    except Exception as e:
        logger.error(f"[save_edited_field] Неожиданная ошибка: {e}")
        await update.message.reply_text(
            "❌ Непредвиденная ошибка. Свяжитесь с администратором.",
            reply_markup=get_manager_menu()
        )
        return

    context.user_data.clear()
    return ConversationHandler.END


async def back_to_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    # Удаляем текущее сообщение (меню редактирования)
    await query.delete_message()

    # Возвращаем в главное меню
    user_id = update.effective_user.id
    if user_id in MANAGERS:
        await context.bot.send_message(
            chat_id=query.message.chat_id,
            text=f"Добро пожаловать, {MANAGERS[user_id]}!",
            reply_markup=get_manager_menu()
        )
    elif user_id in DRIVERS:
        await context.bot.send_message(
            chat_id=query.message.chat_id,
            text=f"Добро пожаловать, {DRIVERS[user_id]}!",
            reply_markup=get_driver_menu()
        )


# Назначение водителя
async def assign_driver(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    delivery_id = int(query.data.split("_")[2])

    # Получаем данные доставки из БД
    cursor = conn.execute(
        """SELECT supplier, pickup_address, delivery_address, cargo_info
           FROM deliveries WHERE id=?""",
        (delivery_id,)
    )
    row = cursor.fetchone()

    if not row:
        await query.edit_message_text("Доставка не найдена.")
        return

    supplier, pickup_addr, delivery_addr, cargo_info = row

    # Формируем текст сообщения с переносами строк
    message_text = (
        f"Выберите водителя для доставки номер {delivery_id}:\n\n"
        f"<b>Поставщик:</b> {supplier}\n"
        f"<b>Адрес загрузки:</b> {pickup_addr or '—'}\n"
        f"<b>Адрес отгрузки:</b> {delivery_addr or '—'}\n"
        f"<b>Комментарий:</b> {cargo_info or '—'}"
    )

    # Получаем список водителей
    drivers = DRIVERS.items()  # (user_id, name)

    if not drivers:
        await query.edit_message_text("Водителей нет.")
        return

    keyboard = []
    for driver_id, name in drivers:
        keyboard.append([
            InlineKeyboardButton(
                name,
                callback_data=f"set_driver_{delivery_id}_{driver_id}"
            )
        ])
    reply_markup = InlineKeyboardMarkup(keyboard)

    # Отправляем сообщение с HTML-разметкой (для жирного текста)
    await query.edit_message_text(
        text=message_text,
        reply_markup=reply_markup,
        parse_mode="HTML"  # Включаем HTML для <b>...</b>
    )

async def set_driver(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    data = query.data.split("_")
    delivery_id = int(data[2])
    driver_id = int(data[3])

    driver_name = DRIVERS.get(driver_id)
    if not driver_name:
        await query.edit_message_text("Ошибка: водитель не найден.")
        return

    cursor = conn.execute(
        """SELECT id, supplier, payer, invoice_number, pickup_address,
                   delivery_address, cargo_info, author_name
           FROM deliveries WHERE id = ?""",
        (delivery_id,)
    )
    row = cursor.fetchone()
    if not row:
        await query.edit_message_text("Ошибка: доставка не найдена.")
        return

    delivery_num = row[0]
    supplier = row[1]
    payer = row[2]
    invoice_number = row[3]
    pickup_addr = row[4]
    delivery_addr = row[5]
    cargo_info = row[6]
    author_name = row[7]

    message_text = (
        f"🚚 <b>Новая доставка №{delivery_num}</b>\n\n"
        f"<b>Поставщик:</b> {supplier}\n"
        f"<b>Плательщик:</b> {payer}\n"
        f"<b>Счёт:</b> {invoice_number}\n"
        f"<b>Адрес загрузки:</b> {pickup_addr or '—'}\n"
        f"<b>Адрес отгрузки:</b> {delivery_addr or '—'}\n"
        f"<b>Габариты/вес/комментарий:</b> {cargo_info or '—'}\n"
        f"<b>Автор заявки:</b> {author_name}\n\n"
    )

    keyboard = [
        [InlineKeyboardButton(
            "Принять в работу", callback_data=f"accept_{delivery_id}"
        )]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    try:
        msg = await context.bot.send_message(
            chat_id=driver_id,
            text=message_text,
            reply_markup=reply_markup,
            parse_mode="HTML"
        )
        
        # Сохраняем ID сообщения в отдельный ключ (чтобы не мешать с bot_messages)
        if 'assigned_delivery_messages' not in context.user_data:
            context.user_data['assigned_delivery_messages'] = []
        context.user_data['assigned_delivery_messages'].append(msg.message_id)

        conn.execute(
            """UPDATE deliveries
               SET driver_id = ?, driver_name = ?, status = 'принята в работу'
               WHERE id = ?""",
            (driver_id, driver_name, delivery_id)
        )
        conn.commit()

        await query.delete_message()
    except Exception as e:
        logger.error(f"Не удалось отправить сообщение водителю {driver_id}: {e}")
        await query.edit_message_text("Произошла ошибка при отправке сообщения водителю.")


# Принятые в работу
async def in_work_deliveries(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    # Удаляем старое меню
    await query.delete_message()
    await delete_excel_message(context, query.message.chat_id)

    # Очищаем предыдущие сообщения
    if 'delivery_messages' in context.user_data:
        for msg_id in context.user_data['delivery_messages']:
            try:
                await context.bot.delete_message(
                    chat_id=query.message.chat_id,
                    message_id=msg_id
                )
            except:
                pass
        context.user_data.pop('delivery_messages', None)

    # Получаем доставки
    cursor = conn.execute(
        """SELECT id, supplier, payer, invoice_number, pickup_address,
                   delivery_address, cargo_info, author_name, driver_name
           FROM deliveries WHERE status='принята в работу'"""
    )
    deliveries = cursor.fetchall()

    if not deliveries:
        msg = await context.bot.send_message(
            chat_id=query.message.chat_id,
            text="Нет доставок в работе.",
            reply_markup=get_manager_menu()
        )
        context.user_data['delivery_messages'] = [msg.message_id]
        return

    # Формируем единый текст
    text = "🚚 Доставки в работе:\n\n"
    for d in deliveries:
        text += (f"№{d[0]} | {d[1]} → {d[5] or '—'}\n"
                 f"   Плательщик: {d[2]} | Счёт: {d[3]}\n"
                 f"   Адрес загрузки: {d[4] or '—'}\n"
                 f"   Водитель: {d[8] or 'Не назначен'}\n\n")

    # Отправляем сообщение БЕЗ кнопок
    msg = await context.bot.send_message(
        chat_id=query.message.chat_id,
        text=text,
        reply_markup=get_manager_menu()  # Возвращаем основное меню
    )

    context.user_data['delivery_messages'] = [msg.message_id]


# Выполненные доставки
async def completed_deliveries(update: Update, context:ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    # Удаляем старое меню
    await query.delete_message()
    await delete_excel_message(context, query.message.chat_id)

    # Очищаем предыдущие сообщения
    if 'delivery_messages' in context.user_data:
        for msg_id in context.user_data['delivery_messages']:
            try:
                await context.bot.delete_message(
                    chat_id=query.message.chat_id,
                    message_id=msg_id
                )
            except:
                pass
        context.user_data.pop('delivery_messages', None)

    # Получаем доставки (последние 10, с сортировкой по дате выполнения)
    cursor = conn.execute(
        """SELECT id, supplier, payer, invoice_number, pickup_address,
                   delivery_address, cargo_info, author_name, completed_at, driver_name
           FROM deliveries 
           WHERE status='доставлено'
           ORDER BY completed_at DESC
           LIMIT 10"""
    )
    deliveries = cursor.fetchall()

    if not deliveries:
        msg = await context.bot.send_message(
            chat_id=query.message.chat_id,
            text="Нет выполненных доставок.",
            reply_markup=get_manager_menu()
        )
        context.user_data['delivery_messages'] = [msg.message_id]
        return

    # Формируем единый текст с указанием водителя
    text = "✅ Выполненные доставки (последние 10):\n\n"
    for d in deliveries:
        completed_time = datetime.strptime(d[8], "%Y-%m-%d %H:%M:%S")
        text += (f"№{d[0]} | {d[1]} → {d[5] or '—'}\n"
                 f"   Плательщик: {d[2]} | Счёт: {d[3]}\n"
                 f"   Дата выполнения: {completed_time.strftime('%d.%m.%Y %H:%M')}\n"
                 f"   Водитель: {d[9] or 'Не назначен'}\n\n")

    # Отправляем сообщение БЕЗ кнопок
    msg = await context.bot.send_message(
        chat_id=query.message.chat_id,
        text=text,
        reply_markup=get_manager_menu()  # Возвращаем основное меню
    )

    context.user_data['delivery_messages'] = [msg.message_id]



async def cancel_search(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.clear()
    await update.message.reply_text(
        "Поиск отменён.",
        reply_markup=get_manager_menu()
    )

# Скачать таблицу
async def download_table(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    # Удаляем старое меню
    await query.delete_message()
    await delete_excel_message(context, query.message.chat_id)

    if 'delivery_messages' in context.user_data:
        for msg_id in context.user_data['delivery_messages']:
            try:
                await context.bot.delete_message(
                    chat_id=query.message.chat_id,
                    message_id=msg_id
                )
            except:
                pass
        context.user_data.pop('delivery_messages', None)

    # Получаем все доставки из БД
    cursor = conn.execute(
        """SELECT id, supplier, payer, invoice_number, pickup_address,
                   delivery_address, cargo_info, author_name, driver_name, status, completed_at
           FROM deliveries"""
    )
    deliveries = cursor.fetchall()

    if not deliveries:
        msg = await context.bot.send_message(
            chat_id=query.message.chat_id,
            text="Нет данных для экспорта.",
            reply_markup=get_manager_menu()
        )
        context.user_data['delivery_messages'] = [msg.message_id]
        return

    # Создаём Excel-файл
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Доставки"

    # Заголовки таблицы (добавили "Водитель")
    headers = [
        "ID", "Поставщик", "Плательщик", "Счёт", "Адрес загрузки",
        "Адрес отгрузки", "Габариты/вес", "Автор", "Водитель", "Статус", "Дата выполнения"
    ]
    ws.append(headers)

    # Заполняем данными
    for delivery in deliveries:
        ws.append([
            delivery[0],  # ID
            delivery[1],  # Поставщик
            delivery[2],  # Плательщик
            delivery[3],  # Счёт
            delivery[4] or "",  # Адрес загрузки
            delivery[5] or "",  # Адрес отгрузки
            delivery[6] or "",  # Габариты/вес
            delivery[7],  # Автор
            delivery[8] or "",  # Водитель (было delivery[8], теперь это driver_name)
            delivery[9],  # Статус
            delivery[10] or ""  # Дата выполнения (было delivery[9], теперь delivery[10])
        ])

    # 2. Включаем автофильтры для всей таблицы (от шапки до последней строки)
    ws.auto_filter.ref = f"A1:K{ws.max_row}"  # K — 11-й столбец (было J → теперь K)

    # 3. Стилизуем шапку (первая строка)
    header_row = ws[1]
    for cell in header_row:
        # Полужирный шрифт
        cell.font = openpyxl.styles.Font(bold=True)
        # Цвет фона (светло-серый)
        cell.fill = openpyxl.styles.PatternFill(
            start_color="D9D9D9",
            end_color="D9D9D9",
            fill_type="solid"
        )
        # Выравнивание по центру
        cell.alignment = openpyxl.styles.Alignment(
            horizontal="center",
            vertical="center"
        )
        # Автоширина столбцов по содержимому
        column_letter = cell.column_letter
        column_width = max(
            len(str(cell.value)) + 2,
            12  # минимальная ширина
        )
        ws.column_dimensions[column_letter].width = column_width

    # 4. Сохраняем файл во временный буфер
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)  # Возвращаем указатель в начало

    # 5. Отправляем файл пользователю
    sent_message = await context.bot.send_document(
        chat_id=query.message.chat_id,
        document=output,
        filename="доставки.xlsx",
        caption="Таблица доставок"
    )

    # Сохраняем ID сообщения с документом
    context.user_data['excel_message_id'] = sent_message.message_id

    # 6. Отправляем сообщение о завершении
    msg = await context.bot.send_message(
        chat_id=query.message.chat_id,
        text="Таблица успешно экспортирована!",
        reply_markup=get_manager_menu()
    )
    context.user_data['delivery_messages'] = [msg.message_id]


async def delete_excel_message(context: ContextTypes.DEFAULT_TYPE, chat_id: int):
    """Удаляет сообщение с Excel-файлом, если оно есть"""
    if 'excel_message_id' in context.user_data:
        try:
            await context.bot.delete_message(
                chat_id=chat_id,
                message_id=context.user_data['excel_message_id']
            )
            # Очищаем ID после удаления
            context.user_data.pop('excel_message_id', None)
        except Exception as e:
            logger.warning(f"Не удалось удалить сообщение с Excel: {e}")

# Действия водителя: принять в работу
async def accept_delivery(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    delivery_id = int(query.data.split("_")[1])
    driver_chat_id = query.message.chat_id

    # 1. Обновляем статус доставки в БД
    try:
        conn.execute(
            "UPDATE deliveries SET status='принята в работу', work_started_at=? WHERE id=?",
            (datetime.now().isoformat(), delivery_id)
        )
        conn.commit()
    except sqlite3.Error as e:
        logger.error(f"Ошибка при обновлении статуса доставки {delivery_id}: {e}")
        await query.edit_message_text("Произошла ошибка. Попробуйте ещё раз.")
        return

    # 2. Удаляем ВСЕ сообщения о доставках из чата водителя
    if 'bot_messages' in context.user_data:
        for msg_id in context.user_data['bot_messages']:
            try:
                await context.bot.delete_message(
                    chat_id=driver_chat_id,
                    message_id=msg_id
                )
            except Exception as e:
                logger.warning(f"Не удалось удалить сообщение {msg_id}: {e}")
        context.user_data.pop('bot_messages', None)  # Очищаем список

    # 3. Вызываем driver_deliveries — он отправит новый список (только принятые/завершённые)
    await driver_deliveries(update, context)

# Действия водителя: доставить
async def delivered_delivery(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    delivery_id = int(query.data.split("_")[1])

    # Формируем сообщение с подтверждением
    msg = (f"Вы уверены, что доставка №{delivery_id} выполнена?\n\n"
            "После подтверждения статус изменится на «доставлено».")

    keyboard = [
        [
            InlineKeyboardButton("Да, доставлено", callback_data=f"confirm_delivered_{delivery_id}_yes"),
            InlineKeyboardButton("Нет, отменить", callback_data=f"confirm_delivered_{delivery_id}_no")
        ]
    ]

    await query.edit_message_text(text=msg, reply_markup=InlineKeyboardMarkup(keyboard))

async def confirm_delivered(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    data = query.data.split("_")
    delivery_id = int(data[2])
    choice = data[3]  # "yes" или "no"


    if choice == "yes":
        # Формируем дату без миллисекунд и с пробелом вместо T
        completed_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        conn.execute(
            "UPDATE deliveries SET status='доставлено', completed_at=? WHERE id=?",
            (completed_at, delivery_id)
        )
        conn.commit()
        # Полностью удаляем сообщение из чата
        try:
            await query.delete_message()
        except Exception as e:
            logger.error(f"Не удалось удалить сообщение: {e}")
            # Если удаление не удалось — можно отправить пустое обновление
            await query.edit_message_text(" ")
    else:
        # Получаем ВСЕ поля доставки из БД
        cursor = conn.execute(
            """SELECT id, supplier, payer, invoice_number, pickup_address,
                       delivery_address, cargo_info, author_name
               FROM deliveries WHERE id = ?""",
            (delivery_id,)
        )
        row = cursor.fetchone()
        if not row:
            await query.edit_message_text("Ошибка: доставка не найдена.")
            return

        # Распаковываем данные доставки
        delivery_num = row[0]
        supplier = row[1]
        payer = row[2]
        invoice_number = row[3]
        pickup_addr = row[4]
        delivery_addr = row[5]
        cargo_info = row[6]
        author_name = row[7]


        # Формируем ПОЛНОЕ сообщение (как при назначении)
        message_text = (
            f"🚚 <b>Доставка №{delivery_num}</b>\n\n"
            f"<b>Поставщик:</b> {supplier}\n"
            f"<b>Плательщик:</b> {payer}\n"
            f"<b>Счёт:</b> {invoice_number}\n"
            f"<b>Адрес загрузки:</b> {pickup_addr or '—'}\n"
            f"<b>Адрес отгрузки:</b> {delivery_addr or '—'}\n"
            f"<b>Габариты/вес/комментарий:</b> {cargo_info or '—'}\n"
            f"<b>Автор заявки:</b> {author_name}\n\n"
            f"<i>Подтвердите выполнение доставки.</i>"
        )

        # Создаём клавиатуру с кнопкой «Доставлено» (а не «Принять в работу»)
        keyboard = [
            [InlineKeyboardButton(
                "Доставлено", callback_data=f"delivered_{delivery_id}_yes"
            )]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)

        # Обновляем сообщение (заменяем текущее на полное с кнопкой «Доставлено»)
        await query.edit_message_text(
            text=message_text,
            reply_markup=reply_markup,
            parse_mode="HTML"
        )


# Просмотр выполненных доставок водителем
async def driver_completed(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    await query.delete_message()

    # Удаляем все предыдущие сообщения бота
    if 'bot_messages' in context.user_data:
        for msg_id in context.user_data['bot_messages']:
            try:
                await context.bot.delete_message(
                    chat_id=query.message.chat_id,
                    message_id=msg_id
                )
            except:
                pass
        context.user_data.pop('bot_messages', None)

    user_id = update.effective_user.id
    limit = context.user_data.get('completed_count', 5)
    one_month_ago = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")

    cursor = conn.execute(
        """SELECT id, supplier, invoice_number, completed_at
           FROM deliveries
           WHERE driver_id=? AND status='доставлено' AND completed_at >= ?
           ORDER BY completed_at DESC
           LIMIT ?""",
        (user_id, one_month_ago, limit)
    )
    deliveries = cursor.fetchall()

    if not deliveries:
        msg = await context.bot.send_message(
            chat_id=query.message.chat_id,
            text="Нет завершённых доставок за последний месяц.",
            reply_markup=get_driver_menu()
        )
        context.user_data['bot_messages'] = [msg.message_id]
        return

    msg_text = f"Завершённые доставки (последние {limit}):\n\n"
    for d in deliveries:
        completed_time = datetime.strptime(d[3], "%Y-%m-%d %H:%M:%S")
        msg_text += (f"ID: {d[0]}\n"
                     f"Поставщик: {d[1]}\n"
                     f"Счёт: {d[2]}\n"
                     f"Завершено: {completed_time.strftime('%d.%m.%Y %H:%M')}\n\n")

    msg = await context.bot.send_message(
        chat_id=query.message.chat_id,
        text=msg_text,
        reply_markup=get_driver_menu()
    )
    context.user_data['bot_messages'] = [msg.message_id]


async def download_table_driver(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    # Удаляем старое меню
    await query.delete_message()
    await delete_excel_message(context, query.message.chat_id)


    if 'delivery_messages' in context.user_data:
        for msg_id in context.user_data['delivery_messages']:
            try:
                await context.bot.delete_message(
                    chat_id=query.message.chat_id,
                    message_id=msg_id
                )
            except:
                pass
        context.user_data.pop('delivery_messages', None)

    driver_id = update.effective_user.id

    # Получаем доставки, назначенные текущему водителю (статус != 'черновик')
    cursor = conn.execute(
        """SELECT id, supplier, payer, invoice_number, pickup_address,
                   delivery_address, cargo_info, author_name, driver_name, status, completed_at
           FROM deliveries
           WHERE driver_id = ? AND status != 'черновик'
           ORDER BY completed_at DESC"""
        , (driver_id,)
    )
    deliveries = cursor.fetchall()


    if not deliveries:
        msg = await context.bot.send_message(
            chat_id=query.message.chat_id,
            text="У вас нет доставок для экспорта.",
            reply_markup=get_driver_menu()
        )
        context.user_data['delivery_messages'] = [msg.message_id]
        return

    # Создаём Excel-файл
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Мои доставки"

    # Заголовки таблицы
    headers = [
        "ID", "Поставщик", "Плательщик", "Счёт", "Адрес загрузки",
        "Адрес отгрузки", "Габариты/вес", "Автор", "Водитель", "Статус", "Дата выполнения"
    ]
    ws.append(headers)

    # Заполняем данными
    for delivery in deliveries:
        ws.append([
            delivery[0],  # ID
            delivery[1],  # Поставщик
            delivery[2],  # Плательщик
            delivery[3],  # Счёт
            delivery[4] or "",  # Адрес загрузки
            delivery[5] or "",  # Адрес отгрузки
            delivery[6] or "",  # Габариты/вес
            delivery[7],  # Автор
            delivery[8],  # Водитель
            delivery[9],  # Статус
            delivery[10] or ""  # Дата выполнения
        ])

    # Включаем автофильтры
    ws.auto_filter.ref = f"A1:K{ws.max_row}"

    # Стилизуем шапку
    header_row = ws[1]
    for cell in header_row:
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(
            start_color="D9D9D9",
            end_color="D9D9D9",
            fill_type="solid"
        )
        cell.alignment = openpyxl.styles.Alignment(
            horizontal="center",
            vertical="center"
        )
        column_letter = cell.column_letter
        column_width = max(len(str(cell.value)) + 2, 12)
        ws.column_dimensions[column_letter].width = column_width


    # Сохраняем файл во временный буфер
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Отправляем файл
    sent_message = await context.bot.send_document(
        chat_id=query.message.chat_id,
        document=output,
        filename="мои_доставки.xlsx",
        caption="Ваши доставки"
    )

    context.user_data['excel_message_id'] = sent_message.message_id


    # Сообщение о завершении
    msg = await context.bot.send_message(
        chat_id=query.message.chat_id,
        text="Таблица успешно экспортирована!",
        reply_markup=get_driver_menu()
    )
    context.user_data['delivery_messages'] = [msg.message_id]


async def back_to_driver_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    # Удаляем текущее сообщение (результат предыдущего действия)
    try:
        await query.delete_message()
    except Exception as e:
        logger.warning(f"Не удалось удалить сообщение: {e}")

    # Отправляем новое сообщение с меню водителя
    user_id = update.effective_user.id
    driver_name = DRIVERS.get(user_id, "Водитель")
    
    await context.bot.send_message(
        chat_id=query.message.chat_id,
        text=f"Здравствуйте, {driver_name}! Выберите действие:",
        reply_markup=get_driver_menu()
    )

# Вспомогательные функции
def get_driver_name(driver_id: int) -> str:
    """Получить имя водителя по ID."""
    return DRIVERS.get(driver_id, "Неизвестен")

def get_manager_name(manager_id: int) -> str:
    """Получить имя менеджера по ID."""
    return MANAGERS.get(manager_id, "Неизвестен")

# Инициализация БД (если таблицы не созданы)
def init_db():
    conn = sqlite3.connect('logistics.db')
    
    # Таблица доставок (обновлённая)
    conn.execute('''
    CREATE TABLE IF NOT EXISTS deliveries (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        supplier TEXT NOT NULL,
        payer TEXT NOT NULL,
        invoice_number TEXT NOT NULL,
        pickup_address TEXT,
        delivery_address TEXT,
        cargo_info TEXT,
        driver_id INTEGER,
        driver_name TEXT,
        author_name TEXT NOT NULL,
        status TEXT DEFAULT 'черновик',
        work_started_at TIMESTAMP,
        completed_at TIMESTAMP,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )
    ''')
    
    conn.commit()
    conn.close()

async def show_main_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if user_id in MANAGERS:
        keyboard = get_manager_menu()
        await update.message.reply_text(
            "Выберите действие:",
            reply_markup=keyboard
        )
    elif user_id in DRIVERS:
        keyboard = get_driver_menu()
        await update.message.reply_text(
            "Выберите действие:",
            reply_markup=keyboard
        )
    else:
        await update.message.reply_text("У вас нет доступа к меню.")

# Основной обработчик
def main():
    application = Application.builder().token("8344348942:AAGTzHRkWE-Yr6uvCf6Mn_Pgj3WCOjigNGI").build()


    # Основной ConversationHandler — создание И РЕДАКТИРОВАНИЕ доставки
    conv_handler = ConversationHandler(
        entry_points=[
            CallbackQueryHandler(add_delivery, pattern="^add_delivery$"),
            CallbackQueryHandler(edit_delivery, pattern="^edit_delivery_.*"),
            CallbackQueryHandler(select_edit_field, pattern=r"^field:.*"),
        ],
        states={
            SUPPLIER: [MessageHandler(filters.TEXT, supplier_step)],
            PAYER: [MessageHandler(filters.TEXT & ~filters.COMMAND, payer_step)],
            INVOICE: [MessageHandler(filters.TEXT & ~filters.COMMAND, invoice_step)],
            PICKUP: [MessageHandler(filters.TEXT & ~filters.COMMAND, pickup_step)],
            DELIVERY: [MessageHandler(filters.TEXT & ~filters.COMMAND, delivery_step)],
            CARGO_INFO: [MessageHandler(filters.TEXT & ~filters.COMMAND, cargo_info_step)],
            EDITING_FIELD: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, save_edited_field)
            ]
        },
        fallbacks=[CommandHandler("cancel", cancel_search)],
        per_message=False
    )

    application.add_handler(conv_handler)

    # Обработчики команд
    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("cancel", cancel_search))

    # CallbackQueryHandler для меню и действий
    application.add_handler(CallbackQueryHandler(download_table, pattern="^download_table$"))
    application.add_handler(CallbackQueryHandler(assign_driver, pattern="^assign_driver_.*"))
    application.add_handler(CallbackQueryHandler(set_driver, pattern="^set_driver_.*"))
    application.add_handler(CallbackQueryHandler(accept_delivery, pattern="^accept_.*"))
    application.add_handler(CallbackQueryHandler(delivered_delivery, pattern="^delivered_.*"))
    application.add_handler(CallbackQueryHandler(confirm_delivered, pattern="^confirm_delivered_.*"))
    application.add_handler(CallbackQueryHandler(planned_deliveries, pattern="^planned_deliveries$"))
    application.add_handler(CallbackQueryHandler(in_work_deliveries, pattern="^in_work_deliveries$"))
    application.add_handler(CallbackQueryHandler(completed_deliveries, pattern="^completed_deliveries$"))
    application.add_handler(CallbackQueryHandler(driver_completed, pattern="^driver_completed$"))
    application.add_handler(CallbackQueryHandler(back_to_driver_menu, pattern="^back_to_driver_menu$"))
    application.add_handler(CallbackQueryHandler(back_to_menu, pattern="^back_to_menu$"))
    application.add_handler(CallbackQueryHandler(download_table_driver, pattern="^download_table_driver$"))
    application.add_handler(CallbackQueryHandler(driver_deliveries, pattern="^driver_deliveries$"))
    # !!! УБРАТЬ эту строку — она больше не нужна
    # application.add_handler(CallbackQueryHandler(select_edit_field, pattern=r"^field:.*"))


    try:
        print("Бот запущен. Ожидание обновлений...")
        application.run_polling(
            poll_interval=3.0,
            timeout=30,
            allowed_updates=Update.ALL_TYPES
        )
    except Exception as e:
        logger.critical(f"Критическая ошибка при запуске бота: {e}", exc_info=True)
    finally:
        if conn:
            conn.close()
            print("Соединение с БД закрыто.")


if __name__ == '__main__':
    init_db()  # Инициализируем БД при запуске
    main()